#!/usr/bin/env python3

import sys
from openpyxl import Workbook #Does not zero index rows and columns

# Set MCNP output filename
tallyfile = sys.argv[1]
flag = "1tally"

def make_list(data, i):
    """
    convert line entry into a list with no extra " " spaces
    if file contains blank line, returns a single space " " list
    """
    row = data[i].rstrip().split(" ")
    line = []
    for elem in row:
        if elem != "":
            line.append(elem)
    if len(line) == 0:
        return [" "]
    else:
        return line
    
def get_cell_data(data, i, FMF):
    """
    given data file and starting i index of line to begin cell block
    outputs list of 3 lists: [[cells],[values],[error],[values*FMF]]
    """
    cells, values, error, FMF_list = [], [], [], []
    for i in range(i,len(data),3):
        if make_list(data,i)[0] == "cell":
            cells.append(float(make_list(data,i)[1]))
            values.append(float(make_list(data,i+1)[0]))
            error.append(float(make_list(data, i+1)[1]))
            FMF_list.append(float(make_list(data,i+1)[0])*FMF)
        else:
            break
    return [cells, values, error, FMF_list]

def find_keff(data):
    """
    parses output file and pulls keff value
    returns keff as string
    """
    for i in range(len(data)):
        line = make_list(data,i)
        if line[0] == "final" and line[1] == "result":
            keff = line[2]
            break
    return keff

def find_wloss(data):
    """
    parses output file and pulls weight loss to fission value
    returns wloss as string
    """
    for i in range(len(data)):
        line = make_list(data,i)
        if line[0] == "prompt" and line[1] == "fission":
            wloss = line[9]
            break
    return wloss

def compute_FMF(data):
    """
    Calculates FMF value using data from output file
    """
    P = 5700000
    keff = float(find_keff(data))
    wloss = float(find_wloss(data))
    qavg = 201.76 #MeV/fsn
    convert = 1.6021773e-13 #J/MeV
    FMF = (P*keff/wloss)/(qavg*convert*keff) # fsn n /sec
    return FMF

def make_str(str_list):
    """
    turns list of strings into combined string
    intended for particle list because I'm too lazy to change the data pull
    """
    delimiter = " "
    return delimiter.join(str_list)
    

def write_tally_data_sheet(tally, xlsx, first_sheet=False):
    """
    tally in the form of a list with entries:
        [tallynum, units, particles, tally_data]
        where:
            tallynum = string
            units = string
            particles = string
            tally data = list of 4 lists with:
                [cells], [values], [error], [values*FMF]
    
    writes the tally data to the given workbook on its own sheet
    """
    if first_sheet == True:
        xlsx.title = "tally_"+str(tally[0])
    else:
        xlsx = wb.create_sheet(title="tally_"+str(tally[0]))
        
    params = ["","","","","","tally","units","type","FMF", "J/MeV"]
    headers = ["cell", "value", "error", "FMF*value", "", tally[0], tally[1], tally[2], FMF, 1.6021773e-13]
    for col in range(1, len(params)+1):
        xlsx.cell(column=col, row=1, value=params[col-1])
    for col in range(1, len(headers)+1):
        xlsx.cell(column=col, row=2, value=headers[col-1])
    tally_data = tally[3]
    for col in range(1, len(tally_data)+1):
        for row in range(3, len(tally_data[0])+3):
            xlsx.cell(column=col, row=row, value=tally_data[col-1][row-3])   

tallies = []
results = open(tallyfile, "r")
data = results.readlines()
FMF = compute_FMF(data)
for i in range(len(data)):
    line = make_list(data, i)
    if line[0] == flag and line[1] != "fluctuation":
        # find cell block start
        for l in range(i,len(data)):
            if make_list(data,l)[0] == "cell":
                j = l
                break
            else:
                continue
        
        #begin processing of tally block
        tallynum = line[1]                    # string
        units = make_list(data, i+1)[-1]      # string
        particles = make_str(make_list(data, i+2)[2:])  # string
        
        # process cell block
        tally_data = get_cell_data(data, j, FMF)
        tallies.append([tallynum, units, particles, tally_data])

# Implementation - Make xlsx file
wb = Workbook()
wb_name = tallyfile + ".xlsx"  

sheet1 = wb.active
write_tally_data_sheet(tallies[0], sheet1, first_sheet=True)

for i in range(1, len(tallies)):
    write_tally_data_sheet(tallies[i], i)

wb.save(filename = wb_name)
      

